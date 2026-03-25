#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel到Markdown转换工具 - 工具函数模块

提供通用的辅助功能，包括文件处理、文本清理、路径管理等。

作者: Battery Cost Monitor System
创建时间: 2025-08-28
"""

import os
import re
import logging
from typing import Optional, List, Tuple, Dict, Any
from pathlib import Path
import unicodedata


class FileManager:
    """文件管理工具类"""
    
    @staticmethod
    def ensure_directory(path: str) -> None:
        """确保目录存在，如果不存在则创建"""
        Path(path).mkdir(parents=True, exist_ok=True)
    
    @staticmethod
    def get_safe_filename(filename: str) -> str:
        """
        生成安全的文件名，移除或替换非法字符
        
        Args:
            filename: 原始文件名
            
        Returns:
            安全的文件名
        """
        # 移除或替换Windows文件名中的非法字符
        illegal_chars = r'[<>:"/\\|?*]'
        safe_name = re.sub(illegal_chars, '_', filename)
        
        # 移除开头和结尾的空格和点
        safe_name = safe_name.strip('. ')
        
        # 限制文件名长度
        if len(safe_name) > 255:
            name, ext = os.path.splitext(safe_name)
            safe_name = name[:255-len(ext)] + ext
            
        return safe_name or "unnamed"
    
    @staticmethod
    def get_unique_filename(directory: str, base_name: str, extension: str = "") -> str:
        """
        生成唯一的文件名，如果文件已存在则添加数字后缀
        
        Args:
            directory: 目录路径
            base_name: 基础文件名
            extension: 文件扩展名
            
        Returns:
            唯一的文件名
        """
        counter = 0
        extension = extension if extension.startswith('.') else f'.{extension}' if extension else ''
        
        while True:
            if counter == 0:
                filename = f"{base_name}{extension}"
            else:
                filename = f"{base_name}_{counter}{extension}"
                
            full_path = os.path.join(directory, filename)
            if not os.path.exists(full_path):
                return filename
            counter += 1


class TextProcessor:
    """文本处理工具类"""
    
    @staticmethod
    def clean_text(text: str) -> str:
        """
        清理文本，移除多余的空白字符和控制字符
        
        Args:
            text: 原始文本
            
        Returns:
            清理后的文本
        """
        if not text:
            return ""
            
        # 统一换行符
        text = text.replace('\r\n', '\n').replace('\r', '\n')
        
        # 移除控制字符（保留换行符和制表符）
        text = ''.join(char for char in text 
                      if unicodedata.category(char) != 'Cc' or char in '\n\t')
        
        # 清理多余的空白字符
        text = re.sub(r'[ \t]+', ' ', text)  # 多个空格/制表符合并为一个空格
        text = re.sub(r'\n\s*\n\s*\n+', '\n\n', text)  # 多个空行合并为两个换行符
        
        return text.strip()
    
    @staticmethod
    def escape_markdown(text: str) -> str:
        """
        转义Markdown特殊字符
        
        Args:
            text: 原始文本
            
        Returns:
            转义后的文本
        """
        if not text:
            return ""
            
        # Markdown特殊字符列表
        special_chars = r'[\*_`\[\]()#+\-\.!\\|~>]'
        
        # 转义特殊字符
        escaped_text = re.sub(special_chars, r'\\\g<0>', text)
        
        return escaped_text
    
    @staticmethod
    def format_table_cell(value: Any) -> str:
        """
        格式化表格单元格内容
        
        Args:
            value: 单元格值
            
        Returns:
            格式化后的字符串
        """
        if value is None:
            return ""
        
        # 转换为字符串
        text = str(value)
        
        # 清理文本
        text = TextProcessor.clean_text(text)
        
        # 移除表格分隔符字符
        text = text.replace('|', '\\|')
        text = text.replace('\n', '<br>')
        
        return text
    
    @staticmethod
    def detect_text_format(cell) -> Dict[str, bool]:
        """
        检测单元格文本格式
        
        Args:
            cell: openpyxl单元格对象
            
        Returns:
            格式信息字典
        """
        formats = {
            'bold': False,
            'italic': False,
            'underline': False,
            'strikethrough': False
        }
        
        if hasattr(cell, 'font') and cell.font:
            font = cell.font
            formats['bold'] = bool(font.bold)
            formats['italic'] = bool(font.italic)
            formats['underline'] = bool(font.underline)
            formats['strikethrough'] = bool(font.strike)
        
        return formats


class MarkdownFormatter:
    """Markdown格式化工具类"""
    
    @staticmethod
    def apply_text_format(text: str, formats: Dict[str, bool]) -> str:
        """
        应用文本格式到Markdown
        
        Args:
            text: 原始文本
            formats: 格式信息
            
        Returns:
            格式化的Markdown文本
        """
        if not text:
            return ""
        
        # 应用格式
        if formats.get('bold'):
            text = f"**{text}**"
        if formats.get('italic'):
            text = f"*{text}*"
        if formats.get('strikethrough'):
            text = f"~~{text}~~"
        if formats.get('underline'):
            # Markdown没有原生下划线支持，使用HTML标签
            text = f"<u>{text}</u>"
        
        return text
    
    @staticmethod
    def create_table_header(headers: List[str]) -> str:
        """
        创建Markdown表格头部
        
        Args:
            headers: 表头列表
            
        Returns:
            Markdown表格头部字符串
        """
        if not headers:
            return ""
        
        # 清理和格式化表头
        clean_headers = [TextProcessor.format_table_cell(h) for h in headers]
        
        # 创建表格头部
        header_row = "| " + " | ".join(clean_headers) + " |"
        separator_row = "| " + " | ".join(["---"] * len(clean_headers)) + " |"
        
        return f"{header_row}\n{separator_row}"
    
    @staticmethod
    def create_table_row(cells: List[Any]) -> str:
        """
        创建Markdown表格行
        
        Args:
            cells: 单元格值列表
            
        Returns:
            Markdown表格行字符串
        """
        clean_cells = [TextProcessor.format_table_cell(cell) for cell in cells]
        return "| " + " | ".join(clean_cells) + " |"
    
    @staticmethod
    def create_heading(text: str, level: int = 1) -> str:
        """
        创建Markdown标题
        
        Args:
            text: 标题文本
            level: 标题级别 (1-6)
            
        Returns:
            Markdown标题字符串
        """
        level = max(1, min(6, level))  # 限制在1-6范围内
        return f"{'#' * level} {text}"
    
    @staticmethod
    def create_image_link(alt_text: str, image_path: str, title: str = "") -> str:
        """
        创建Markdown图片链接
        
        Args:
            alt_text: 替代文本
            image_path: 图片路径
            title: 图片标题（可选）
            
        Returns:
            Markdown图片链接字符串
        """
        if title:
            return f"![{alt_text}]({image_path} \"{title}\")"
        else:
            return f"![{alt_text}]({image_path})"


class LogManager:
    """日志管理工具类"""
    
    @staticmethod
    def setup_logger(name: str, log_file: Optional[str] = None, level: int = logging.INFO) -> logging.Logger:
        """
        设置日志记录器
        
        Args:
            name: 日志记录器名称
            log_file: 日志文件路径（可选）
            level: 日志级别
            
        Returns:
            配置好的日志记录器
        """
        logger = logging.getLogger(name)
        logger.setLevel(level)
        
        # 清除现有的处理器
        logger.handlers.clear()
        
        # 创建格式化器
        formatter = logging.Formatter(
            '%(asctime)s - %(name)s - %(levelname)s - %(message)s',
            datefmt='%Y-%m-%d %H:%M:%S'
        )
        
        # 控制台处理器
        console_handler = logging.StreamHandler()
        console_handler.setLevel(level)
        console_handler.setFormatter(formatter)
        logger.addHandler(console_handler)
        
        # 文件处理器（如果指定了日志文件）
        if log_file:
            FileManager.ensure_directory(os.path.dirname(log_file))
            file_handler = logging.FileHandler(log_file, encoding='utf-8')
            file_handler.setLevel(level)
            file_handler.setFormatter(formatter)
            logger.addHandler(file_handler)
        
        return logger


class ExcelHelper:
    """Excel处理辅助工具类"""
    
    @staticmethod
    def get_cell_value(cell) -> Any:
        """
        获取单元格的值，处理公式和特殊情况
        
        Args:
            cell: openpyxl单元格对象
            
        Returns:
            单元格的值
        """
        if cell is None:
            return None
        
        # 如果是公式，返回计算结果
        if hasattr(cell, 'data_type') and cell.data_type == 'f':
            return cell.value if cell.value is not None else ""
        
        return cell.value
    
    @staticmethod
    def is_merged_cell(worksheet, row: int, col: int) -> Tuple[bool, Optional[Tuple[int, int, int, int]]]:
        """
        检查单元格是否是合并单元格的一部分
        
        Args:
            worksheet: 工作表对象
            row: 行号（1-based）
            col: 列号（1-based）
            
        Returns:
            (是否合并, 合并范围) - 合并范围格式：(min_row, min_col, max_row, max_col)
        """
        for merged_range in worksheet.merged_cells.ranges:
            if (merged_range.min_row <= row <= merged_range.max_row and 
                merged_range.min_col <= col <= merged_range.max_col):
                return True, (merged_range.min_row, merged_range.min_col, 
                            merged_range.max_row, merged_range.max_col)
        
        return False, None
    
    @staticmethod
    def get_used_range(worksheet) -> Tuple[int, int, int, int]:
        """
        获取工作表的使用范围
        
        Args:
            worksheet: 工作表对象
            
        Returns:
            (min_row, min_col, max_row, max_col) 使用范围
        """
        if worksheet.max_row == 1 and worksheet.max_column == 1:
            # 检查是否真的有数据
            if worksheet.cell(1, 1).value is None:
                return 0, 0, 0, 0
        
        return (worksheet.min_row or 1, 
                worksheet.min_column or 1, 
                worksheet.max_row or 1, 
                worksheet.max_column or 1)


# 导出的公共函数
def validate_excel_file(file_path: str) -> bool:
    """
    验证Excel文件是否有效
    
    Args:
        file_path: Excel文件路径
        
    Returns:
        文件是否有效
    """
    if not os.path.exists(file_path):
        return False
    
    try:
        import openpyxl
        openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        return True
    except Exception:
        return False


def get_output_paths(input_file: str, output_dir: str = None) -> Dict[str, str]:
    """
    生成输出路径
    
    Args:
        input_file: 输入Excel文件路径
        output_dir: 输出目录（可选）
        
    Returns:
        输出路径字典
    """
    input_path = Path(input_file)
    base_name = input_path.stem
    
    if output_dir is None:
        output_dir = input_path.parent / f"{base_name}_markdown"
    else:
        output_dir = Path(output_dir)
    
    FileManager.ensure_directory(str(output_dir))
    
    return {
        'base_dir': str(output_dir),
        'markdown_file': str(output_dir / f"{base_name}.md"),
        'images_dir': str(output_dir / "images"),
        'log_file': str(output_dir / "conversion.log")
    }