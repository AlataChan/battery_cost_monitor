#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel到Markdown转换工具 - Excel解析器模块

负责读取和解析Excel文件，提取文本、表格、图片等内容。

作者: Battery Cost Monitor System
创建时间: 2025-08-28
"""

import os
import logging
from typing import List, Dict, Any, Tuple, Optional, Union
from dataclasses import dataclass
import openpyxl
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.chart import BarChart, LineChart, PieChart, ScatterChart, AreaChart
from PIL import Image
import io

from .utils import (
    ExcelHelper, TextProcessor, LogManager,
    FileManager, get_output_paths
)


@dataclass
class CellInfo:
    """单元格信息"""
    row: int
    col: int
    value: Any
    formats: Dict[str, bool]
    is_merged: bool
    merged_range: Optional[Tuple[int, int, int, int]] = None
    hyperlink: Optional[str] = None


@dataclass
class TableInfo:
    """表格信息"""
    start_row: int
    start_col: int
    end_row: int
    end_col: int
    headers: List[str]
    data: List[List[Any]]
    merged_cells: List[Tuple[int, int, int, int]]


@dataclass
class ImageInfo:
    """图片信息"""
    anchor: str
    title: str
    description: str
    image_data: bytes
    format: str
    size: Tuple[int, int]


@dataclass
class ChartInfo:
    """图表信息"""
    title: str
    chart_type: str
    anchor: str
    size: Tuple[int, int]


@dataclass
class WorksheetInfo:
    """工作表信息"""
    name: str
    tables: List[TableInfo]
    images: List[ImageInfo]
    charts: List[ChartInfo]
    cells: List[CellInfo]
    used_range: Tuple[int, int, int, int]


class ExcelParser:
    """Excel文件解析器"""
    
    def __init__(self, file_path: str, output_dir: str = None):
        """
        初始化Excel解析器
        
        Args:
            file_path: Excel文件路径
            output_dir: 输出目录
        """
        self.file_path = file_path
        self.output_paths = get_output_paths(file_path, output_dir)
        self.logger = LogManager.setup_logger(
            "ExcelParser", 
            self.output_paths['log_file']
        )
        
        self.workbook = None
        self.worksheets_info: List[WorksheetInfo] = []
        
    def load_workbook(self) -> bool:
        """
        加载Excel工作簿
        
        Returns:
            是否成功加载
        """
        try:
            self.logger.info(f"正在加载Excel文件: {self.file_path}")
            self.workbook = openpyxl.load_workbook(
                self.file_path, 
                data_only=False,  # 保留公式
                keep_links=True   # 保留链接
            )
            self.logger.info(f"成功加载工作簿，包含 {len(self.workbook.worksheets)} 个工作表")
            return True
            
        except Exception as e:
            self.logger.error(f"加载Excel文件失败: {str(e)}")
            return False
    
    def parse_all_worksheets(self) -> List[WorksheetInfo]:
        """
        解析所有工作表
        
        Returns:
            工作表信息列表
        """
        if not self.workbook:
            if not self.load_workbook():
                return []
        
        self.worksheets_info = []
        
        for worksheet in self.workbook.worksheets:
            self.logger.info(f"正在解析工作表: {worksheet.title}")
            worksheet_info = self._parse_worksheet(worksheet)
            self.worksheets_info.append(worksheet_info)
        
        self.logger.info(f"完成解析，共处理 {len(self.worksheets_info)} 个工作表")
        return self.worksheets_info
    
    def _parse_worksheet(self, worksheet) -> WorksheetInfo:
        """
        解析单个工作表
        
        Args:
            worksheet: openpyxl工作表对象
            
        Returns:
            工作表信息
        """
        # 获取使用范围
        used_range = ExcelHelper.get_used_range(worksheet)
        
        # 解析单元格
        cells = self._parse_cells(worksheet, used_range)
        
        # 检测和解析表格
        tables = self._detect_tables(worksheet, cells, used_range)
        
        # 解析图片
        images = self._parse_images(worksheet)
        
        # 解析图表
        charts = self._parse_charts(worksheet)
        
        return WorksheetInfo(
            name=worksheet.title,
            tables=tables,
            images=images,
            charts=charts,
            cells=cells,
            used_range=used_range
        )
    
    def _parse_cells(self, worksheet, used_range: Tuple[int, int, int, int]) -> List[CellInfo]:
        """
        解析工作表中的单元格
        
        Args:
            worksheet: 工作表对象
            used_range: 使用范围
            
        Returns:
            单元格信息列表
        """
        cells = []
        min_row, min_col, max_row, max_col = used_range
        
        if max_row == 0:  # 空工作表
            return cells
        
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                cell = worksheet.cell(row, col)
                
                # 获取单元格值
                value = ExcelHelper.get_cell_value(cell)
                
                # 跳过空单元格（除非它是合并单元格的一部分）
                is_merged, merged_range = ExcelHelper.is_merged_cell(worksheet, row, col)
                if value is None and not is_merged:
                    continue
                
                # 获取格式信息
                formats = TextProcessor.detect_text_format(cell)
                
                # 获取超链接
                hyperlink = None
                if hasattr(cell, 'hyperlink') and cell.hyperlink:
                    hyperlink = cell.hyperlink.target
                
                cell_info = CellInfo(
                    row=row,
                    col=col,
                    value=value,
                    formats=formats,
                    is_merged=is_merged,
                    merged_range=merged_range,
                    hyperlink=hyperlink
                )
                
                cells.append(cell_info)
        
        return cells
    
    def _detect_tables(self, worksheet, cells: List[CellInfo], 
                      used_range: Tuple[int, int, int, int]) -> List[TableInfo]:
        """
        检测和解析表格结构
        
        Args:
            worksheet: 工作表对象
            cells: 单元格信息列表
            used_range: 使用范围
            
        Returns:
            表格信息列表
        """
        tables = []
        min_row, min_col, max_row, max_col = used_range
        
        if max_row == 0:
            return tables
        
        # 创建单元格映射
        cell_map = {}
        for cell in cells:
            cell_map[(cell.row, cell.col)] = cell
        
        # 简单的表格检测：连续的非空行作为一个表格
        current_table = None
        
        for row in range(min_row, max_row + 1):
            row_has_data = any((row, col) in cell_map for col in range(min_col, max_col + 1))
            
            if row_has_data:
                if current_table is None:
                    # 开始新表格
                    current_table = {
                        'start_row': row,
                        'start_col': min_col,
                        'end_row': row,
                        'end_col': max_col,
                        'rows': []
                    }
                
                # 添加当前行到表格
                current_table['end_row'] = row
                table_row = []
                for col in range(min_col, max_col + 1):
                    cell = cell_map.get((row, col))
                    table_row.append(cell.value if cell else None)
                current_table['rows'].append(table_row)
                
            else:
                # 空行，结束当前表格
                if current_table and len(current_table['rows']) > 1:
                    table_info = self._create_table_info(current_table, worksheet)
                    tables.append(table_info)
                current_table = None
        
        # 处理最后一个表格
        if current_table and len(current_table['rows']) > 1:
            table_info = self._create_table_info(current_table, worksheet)
            tables.append(table_info)
        
        return tables
    
    def _create_table_info(self, table_data: Dict, worksheet) -> TableInfo:
        """
        创建表格信息对象
        
        Args:
            table_data: 表格数据字典
            worksheet: 工作表对象
            
        Returns:
            表格信息对象
        """
        rows = table_data['rows']
        
        # 第一行作为表头
        headers = [str(cell) if cell is not None else "" for cell in rows[0]]
        
        # 剩余行作为数据
        data = []
        for row in rows[1:]:
            data_row = [cell if cell is not None else "" for cell in row]
            data.append(data_row)
        
        # 获取合并单元格信息
        merged_cells = []
        for merged_range in worksheet.merged_cells.ranges:
            if (table_data['start_row'] <= merged_range.min_row <= table_data['end_row'] and
                table_data['start_col'] <= merged_range.min_col <= table_data['end_col']):
                merged_cells.append((
                    merged_range.min_row, merged_range.min_col,
                    merged_range.max_row, merged_range.max_col
                ))
        
        return TableInfo(
            start_row=table_data['start_row'],
            start_col=table_data['start_col'],
            end_row=table_data['end_row'],
            end_col=table_data['end_col'],
            headers=headers,
            data=data,
            merged_cells=merged_cells
        )
    
    def _parse_images(self, worksheet) -> List[ImageInfo]:
        """
        解析工作表中的图片
        
        Args:
            worksheet: 工作表对象
            
        Returns:
            图片信息列表
        """
        images = []
        
        if not hasattr(worksheet, '_images') or not worksheet._images:
            return images
        
        for image in worksheet._images:
            try:
                # 获取图片信息
                anchor = getattr(image, 'anchor', '')
                title = getattr(image, 'title', '')
                description = getattr(image, 'description', '')
                
                # 获取图片数据
                if hasattr(image, '_data') and image._data:
                    image_data = image._data()
                elif hasattr(image, 'ref') and hasattr(image.ref, '_data'):
                    image_data = image.ref._data()
                else:
                    continue
                
                # 获取图片格式和尺寸
                try:
                    pil_image = Image.open(io.BytesIO(image_data))
                    image_format = pil_image.format.lower() if pil_image.format else 'png'
                    size = pil_image.size
                except Exception:
                    image_format = 'png'
                    size = (0, 0)
                
                image_info = ImageInfo(
                    anchor=str(anchor),
                    title=title,
                    description=description,
                    image_data=image_data,
                    format=image_format,
                    size=size
                )
                
                images.append(image_info)
                
            except Exception as e:
                self.logger.warning(f"解析图片时出错: {str(e)}")
                continue
        
        return images
    
    def _parse_charts(self, worksheet) -> List[ChartInfo]:
        """
        解析工作表中的图表
        
        Args:
            worksheet: 工作表对象
            
        Returns:
            图表信息列表
        """
        charts = []
        
        if not hasattr(worksheet, '_charts') or not worksheet._charts:
            return charts
        
        for chart in worksheet._charts:
            try:
                # 确定图表类型
                chart_type = "unknown"
                if isinstance(chart, BarChart):
                    chart_type = "bar"
                elif isinstance(chart, LineChart):
                    chart_type = "line"
                elif isinstance(chart, PieChart):
                    chart_type = "pie"
                elif isinstance(chart, ScatterChart):
                    chart_type = "scatter"
                elif isinstance(chart, AreaChart):
                    chart_type = "area"
                
                # 获取图表信息
                title = getattr(chart, 'title', '')
                if hasattr(title, 'text') and title.text:
                    title_text = title.text
                else:
                    title_text = f"{chart_type.title()} Chart"
                
                anchor = getattr(chart, 'anchor', '')
                
                # 获取图表尺寸
                width = getattr(chart, 'width', 0)
                height = getattr(chart, 'height', 0)
                size = (int(width), int(height))
                
                chart_info = ChartInfo(
                    title=title_text,
                    chart_type=chart_type,
                    anchor=str(anchor),
                    size=size
                )
                
                charts.append(chart_info)
                
            except Exception as e:
                self.logger.warning(f"解析图表时出错: {str(e)}")
                continue
        
        return charts
    
    def save_images(self, worksheet_info: WorksheetInfo) -> Dict[str, str]:
        """
        保存图片到文件
        
        Args:
            worksheet_info: 工作表信息
            
        Returns:
            图片文件路径映射
        """
        image_paths = {}
        images_dir = self.output_paths['images_dir']
        
        if not worksheet_info.images:
            return image_paths
        
        FileManager.ensure_directory(images_dir)
        
        for i, image_info in enumerate(worksheet_info.images):
            try:
                # 生成文件名
                base_name = f"{worksheet_info.name}_image_{i+1}"
                if image_info.title:
                    base_name = f"{worksheet_info.name}_{FileManager.get_safe_filename(image_info.title)}"
                
                filename = FileManager.get_unique_filename(
                    images_dir, base_name, image_info.format
                )
                
                file_path = os.path.join(images_dir, filename)
                
                # 保存图片
                with open(file_path, 'wb') as f:
                    f.write(image_info.image_data)
                
                # 使用相对路径
                relative_path = f"images/{filename}"
                image_paths[f"image_{i}"] = relative_path
                
                self.logger.info(f"保存图片: {filename}")
                
            except Exception as e:
                self.logger.error(f"保存图片失败: {str(e)}")
                continue
        
        return image_paths
    
    def get_worksheet_by_name(self, name: str) -> Optional[WorksheetInfo]:
        """
        根据名称获取工作表信息
        
        Args:
            name: 工作表名称
            
        Returns:
            工作表信息或None
        """
        for worksheet_info in self.worksheets_info:
            if worksheet_info.name == name:
                return worksheet_info
        return None
    
    def close(self):
        """关闭工作簿"""
        if self.workbook:
            self.workbook.close()
            self.workbook = None